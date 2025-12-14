"""
Report generation module for GovAI Electricity Theft Detection System.
Generates PDF and Excel reports with analysis results.
"""

import os
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
import numpy as np

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference


class ReportGenerator:
    """
    Generates PDF and Excel reports for anomaly detection results.
    """
    
    def __init__(self, output_dir: str = './reports'):
        """
        Initialize report generator.
        
        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Set up custom paragraph styles for PDF."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e3a5f')
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#1e3a5f')
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=12,
            alignment=TA_LEFT
        ))
        
        self.styles.add(ParagraphStyle(
            name='Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        ))
    
    def generate_pdf_report(
        self,
        summary: Dict[str, Any],
        consumers: List[Dict[str, Any]],
        analysis_date: Optional[datetime] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Generate PDF report with analysis results.
        
        Args:
            summary: Analysis summary dictionary
            consumers: List of consumer results
            analysis_date: Date of analysis
            filename: Output filename (without extension)
            
        Returns:
            Path to generated PDF file
        """
        if analysis_date is None:
            analysis_date = datetime.now()
        
        if filename is None:
            filename = f"govai_report_{analysis_date.strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.pdf")
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(
            "GovAI: Electricity Theft Detection Report",
            self.styles['CustomTitle']
        ))
        
        elements.append(Paragraph(
            f"Analysis Date: {analysis_date.strftime('%B %d, %Y at %H:%M')}",
            self.styles['CustomBody']
        ))
        
        elements.append(Spacer(1, 20))
        
        # Executive Summary
        elements.append(Paragraph("Executive Summary", self.styles['CustomHeading']))
        
        summary_text = f"""
        This report presents the results of unsupervised machine learning analysis 
        on electricity consumption data. The analysis identified potential cases of 
        electricity theft or irregular consumption patterns.
        """
        elements.append(Paragraph(summary_text, self.styles['CustomBody']))
        
        elements.append(Spacer(1, 15))
        
        # Summary Statistics Table
        elements.append(Paragraph("Key Findings", self.styles['CustomHeading']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Consumers Analyzed', str(summary.get('total_consumers', 'N/A'))],
            ['Anomalies Detected', str(summary.get('anomalies_detected', 'N/A'))],
            ['High Risk Cases', str(summary.get('high_risk_count', 'N/A'))],
            ['Suspicious Cases', str(summary.get('suspicious_count', 'N/A'))],
            ['Average Anomaly Score', f"{summary.get('avg_anomaly_score', 0):.3f}"],
            ['High Risk Zones', ', '.join(summary.get('high_risk_zones', [])[:3]) or 'None']
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 25))
        
        # Suspicious Consumers Table
        elements.append(Paragraph("Flagged Consumers (Top 20)", self.styles['CustomHeading']))
        
        # Sort consumers by anomaly score and take top 20
        sorted_consumers = sorted(
            consumers, 
            key=lambda x: x.get('anomaly_score', 0), 
            reverse=True
        )[:20]
        
        consumer_data = [['Consumer ID', 'Region', 'Avg Usage', 'Score', 'Status']]
        
        for consumer in sorted_consumers:
            consumer_data.append([
                str(consumer.get('consumer_id', 'N/A')),
                str(consumer.get('region', 'N/A')),
                f"{consumer.get('avg_consumption', 0):.2f}",
                f"{consumer.get('anomaly_score', 0):.3f}",
                str(consumer.get('status', 'N/A'))
            ])
        
        consumer_table = Table(consumer_data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 1.2*inch])
        consumer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        # Color code rows based on status
        for i, consumer in enumerate(sorted_consumers, start=1):
            status = consumer.get('status', '')
            if status == 'High Risk':
                consumer_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fee2e2'))
                ]))
            elif status == 'Suspicious':
                consumer_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fef3c7'))
                ]))
        
        elements.append(consumer_table)
        
        # Footer
        elements.append(Spacer(1, 40))
        elements.append(Paragraph(
            "Report generated by GovAI – Unsupervised Electricity Theft Detection System",
            self.styles['Footer']
        ))
        elements.append(Paragraph(
            "This report is confidential and intended for authorized personnel only.",
            self.styles['Footer']
        ))
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    
    def generate_excel_report(
        self,
        summary: Dict[str, Any],
        consumers: List[Dict[str, Any]],
        analysis_date: Optional[datetime] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Generate Excel report with analysis results.
        
        Args:
            summary: Analysis summary dictionary
            consumers: List of consumer results
            analysis_date: Date of analysis
            filename: Output filename (without extension)
            
        Returns:
            Path to generated Excel file
        """
        if analysis_date is None:
            analysis_date = datetime.now()
        
        if filename is None:
            filename = f"govai_report_{analysis_date.strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        high_risk_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        suspicious_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = "GovAI Electricity Theft Detection Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f"Generated: {analysis_date.strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws_summary['A4'] = "Key Metrics"
        ws_summary['A4'].font = Font(bold=True, size=12)
        
        summary_rows = [
            ('Total Consumers', summary.get('total_consumers', 'N/A')),
            ('Anomalies Detected', summary.get('anomalies_detected', 'N/A')),
            ('High Risk Cases', summary.get('high_risk_count', 'N/A')),
            ('Suspicious Cases', summary.get('suspicious_count', 'N/A')),
            ('Average Anomaly Score', f"{summary.get('avg_anomaly_score', 0):.3f}"),
            ('High Risk Zones', ', '.join(summary.get('high_risk_zones', [])[:5]) or 'None')
        ]
        
        for i, (metric, value) in enumerate(summary_rows, start=5):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].border = thin_border
            ws_summary[f'B{i}'].border = thin_border
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        
        # Consumers Sheet
        ws_consumers = wb.create_sheet("Flagged Consumers")
        
        headers = ['Consumer ID', 'Region', 'Avg Consumption', 'Anomaly Score', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws_consumers.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Sort by anomaly score
        sorted_consumers = sorted(
            consumers,
            key=lambda x: x.get('anomaly_score', 0),
            reverse=True
        )
        
        for row, consumer in enumerate(sorted_consumers, start=2):
            ws_consumers.cell(row=row, column=1, value=consumer.get('consumer_id', 'N/A'))
            ws_consumers.cell(row=row, column=2, value=consumer.get('region', 'N/A'))
            ws_consumers.cell(row=row, column=3, value=round(consumer.get('avg_consumption', 0), 3))
            ws_consumers.cell(row=row, column=4, value=round(consumer.get('anomaly_score', 0), 4))
            ws_consumers.cell(row=row, column=5, value=consumer.get('status', 'N/A'))
            
            # Apply borders and conditional formatting
            for col in range(1, 6):
                cell = ws_consumers.cell(row=row, column=col)
                cell.border = thin_border
                
                status = consumer.get('status', '')
                if status == 'High Risk':
                    cell.fill = high_risk_fill
                elif status == 'Suspicious':
                    cell.fill = suspicious_fill
        
        # Adjust column widths
        ws_consumers.column_dimensions['A'].width = 20
        ws_consumers.column_dimensions['B'].width = 12
        ws_consumers.column_dimensions['C'].width = 18
        ws_consumers.column_dimensions['D'].width = 15
        ws_consumers.column_dimensions['E'].width = 15
        
        # All Consumers Sheet (full data)
        ws_all = wb.create_sheet("All Consumers")
        
        for col, header in enumerate(headers, start=1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row, consumer in enumerate(consumers, start=2):
            ws_all.cell(row=row, column=1, value=consumer.get('consumer_id', 'N/A'))
            ws_all.cell(row=row, column=2, value=consumer.get('region', 'N/A'))
            ws_all.cell(row=row, column=3, value=round(consumer.get('avg_consumption', 0), 3))
            ws_all.cell(row=row, column=4, value=round(consumer.get('anomaly_score', 0), 4))
            ws_all.cell(row=row, column=5, value=consumer.get('status', 'N/A'))
        
        wb.save(filepath)
        
        return filepath


def generate_report(
    summary: Dict[str, Any],
    consumers_df: pd.DataFrame,
    format: str = 'pdf',
    output_dir: str = './reports'
) -> str:
    """
    Convenience function to generate report.
    
    Args:
        summary: Analysis summary
        consumers_df: Consumer results DataFrame
        format: Report format ('pdf' or 'excel')
        output_dir: Output directory
        
    Returns:
        Path to generated report
    """
    generator = ReportGenerator(output_dir)
    
    # Convert DataFrame to list of dicts
    consumers = consumers_df.to_dict('records')
    
    if format.lower() == 'pdf':
        return generator.generate_pdf_report(summary, consumers)
    else:
        return generator.generate_excel_report(summary, consumers)
